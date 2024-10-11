import requests
import os
import pickle
from openpyxl import load_workbook
from openpyxl import Workbook
from lxml import etree as et
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Border, Side
#from openpyxl.workbook import Workbook

if os.path.isfile("program_data"):
    city_list_in=open("program_data", "rb") #nacitanie miest na automaticky scraping
    city_list=pickle.load(city_list_in)
else: #vytvorenie zoznamu miest na auto scraping ak predtym neexistoval (aspon krajske mesta)
    city_list = ['Bratislava', 'Trnava', 'Nitra', 'Trenčín', 'Žilina', 'Prešov', 'Banská Bystrica', 'Košice'] 
    outputfile = open("program_data", "wb")
    pickle.dump(city_list, outputfile)
    outputfile.close()

header = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.66 Safari/537.36"} #hlavičky pre prehliadač
pages_url=[] #zoznam URL pre jednotlivé stránky webu
listing_url=[] #zoznam URL jednotlivých inzerátov
 
#funkcia na prehladavanie stranok a vytvorenie vysledneho suboru s datami   
def scraper(base_url, page_num, file_name, auto_mode: bool):
    pages_url.clear()
    listing_url.clear()
    
    if page_num > 1:
        for i in range (1,page_num+1): 
            page_url=base_url + "?p[page]=" + str(i)
            pages_url.append(page_url)
    else:
        pages_url.append(base_url)
        
    for page in pages_url:
        get_listing_url(page)

    current_directory = os.getcwd()
    file_path = os.path.join(current_directory, file_name) #vytvorenie suboru na zapisovanie udajov
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Vsetky data"
    else:
        wb = load_workbook(file_path)
        ws = wb["Vsetky data"]
        ws.append([]) #medzera pri dopisovani do existujuceho suboru
            
    with open(file_path, 'a', newline='') as file:
        heading=['Link', 'Mesto', 'Ulica', 'Cena', 'Izbovitosť', 'Prenájom/predaj', 'Stav', 'Úžitková plocha v m2', 'Zastavaná plocha v m2', 'EUR/m2']
        ws.append(heading)
        
        use_area_list=[] #zoznamy na vypocet pre novostavby
        eurm2_list=[]
        new_count = 0
        
        other_state=[] #zoznam na inzeraty, ktore nie su novostavby
        other_use_area_list=[]
        other_eurm2_list=[]
        other_count = 0
             
        for list_url in listing_url: 
            listing_dom=get_dom(list_url)
            
            print(list_url)
            
            price=get_price(listing_dom)
            try:
                price=int(''.join(filter(str.isdigit, price)))
                if price==1 or isinstance(price, str): 
                    price = "X"
            except:
                price = "X"
                
            if price=="X":   #Osetrenie aby nezadana cena nekazila data 
                print("\n")
                continue
            
            usable_area=get_usable_area(listing_dom)
            try:
                if "," in usable_area:
                    usable_area = float(usable_area.replace(',', '.').replace('m', ''))
                else:
                    usable_area=int(''.join(filter(str.isdigit, usable_area)))
            except:
                usable_area=="X"
            
            if usable_area=="X": #Osetrenie aby nezadana plocha bytu nekazila data
                print("\n")
                continue
            
            city=get_city(listing_dom)
            street=get_street(listing_dom)
            rooms=get_rooms(listing_dom)
            sale_rent=get_sale_rent(listing_dom)
            state=get_state(listing_dom)
            
            land_area=get_land_area(listing_dom)
            try:
                if "," in land_area:
                    land_area = float(land_area.replace(',', '.').replace('m', ''))
                else:
                    land_area=int(''.join(filter(str.isdigit, land_area)))
            except:
                land_area=="PRIESTOR POZEMKU NIE JE DOSTUPNY"
                       
            try:
                eurm2=price/usable_area  
            except:
                eurm2='CENA/PLOCHA V M2 NIE JE DOSTUPNA'
             
            if state=="Novostavba":
                information =[list_url, city, street, price, rooms, sale_rent, state, usable_area, land_area, eurm2]
                ws.append(information) #vypisovanie udajov inzeratov s novostavbami
                new_count += 1
            else:
                other_state.append([list_url, city, street, price, rooms, sale_rent, state, usable_area, land_area, eurm2])
                other_count += 1
                
            if isinstance(usable_area, (int, float)) and isinstance(eurm2, (int, float)) and state=="Novostavba": #Osetrenie, aby sa vypocet vazeneho priemeru robil len vtedy, ak su oba udaje pritomne (uzitkova plocha a EUR/m2)
                use_area_list.append(usable_area)
                eurm2_list.append(eurm2)
            elif isinstance(usable_area, (int, float)) and isinstance(eurm2, (int, float)) and state!="Novostavba":
                other_use_area_list.append(usable_area)
                other_eurm2_list.append(eurm2)    
                
            print('\n') #odriadkovanie pre krajsi vypis do konzoly
          
        try:  
            new_weigh_avg = sum(x * y for x, y in zip(use_area_list, eurm2_list))/sum(use_area_list) #priemer pre novostavby
            info=['Vazeny priemer', new_weigh_avg, 'Pocet prezrenych inzeratov', new_count] 
            ws.append(info)
            ws.append([])
        except Exception as e:
            new_weigh_avg = None
            print("NIE SU NOVOSTAVBY")
        
        for insertion in other_state: #vypisanie udajov inzeratov, ktore nie su novostavby
            ws.append(insertion)
        
        try:  
            other_weigh_avg = sum(x * y for x, y in zip(other_use_area_list, other_eurm2_list))/sum(other_use_area_list) #priemer pre ostatne
            info=['Vazeny_ priemer', other_weigh_avg, 'Pocet_prezrenych inzeratov', other_count]
            ws.append(info)
            ws.append([])
        except Exception as e:
            other_weigh_avg = None
            print("NIE SU OSTATNE")
        
        wb.save(file_path)
        
        if auto_mode:
            if (new_weigh_avg is not None) and (other_weigh_avg is not None):
                avg_new_other = (new_weigh_avg, other_weigh_avg)
            elif (new_weigh_avg is not None) and (other_weigh_avg is None):
                avg_new_other = (new_weigh_avg, "X")
            elif (new_weigh_avg is None) and (other_weigh_avg is not None):
                avg_new_other = ("X", other_weigh_avg)
                
            return avg_new_other
            
        
def write_summary(file_path, city_detail_list):
    wb = load_workbook(file_path)
    ws_suhrn = wb.create_sheet("Suhrn", 0)
    ws_suhrn.title = "Suhrn"
    
    for city_detail in city_detail_list:
        tags = [city_detail["city"], "1 Izbove byty", "2 Izbove byty", "3 Izbove byty"]
        ws_suhrn.append(tags)
        new_details = [city_detail["insertion_type"]+" novostavby", city_detail["avgs"][0][0], city_detail["avgs"][1][0], city_detail["avgs"][2][0]]
        ws_suhrn.append(new_details)
        other_details = [city_detail["insertion_type"]+" ostatne", city_detail["avgs"][0][1], city_detail["avgs"][1][1], city_detail["avgs"][2][1]]
        ws_suhrn.append(other_details)
        ws_suhrn.append([])
        ws_suhrn.append([])
    
    style_summary(file_path)
    
    wb.save(file_path)
    
def style_summary(file_path):
    wb = load_workbook(file_path)
    ws_suhrn = wb["Suhrn"]

    border = Border(left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'))

    for row in ws_suhrn.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = border
        
            if isinstance(cell.value, str) and cell.value!="X":
                if cell.value.find("novostavby")!=-1 or cell.value.find("ostatne")!=-1:
                    continue
                else:
                    cell.font = Font(bold=True)
            
#funkcia na vrátenie DOM ked zadám URL  
def get_dom(the_url):
    try:
        response = requests.get(the_url, headers=header, timeout=5)
    except requests.exceptions.Timeout as e:
        print(e)
        return get_dom(the_url)
        
    soup = BeautifulSoup(response.text,'lxml')
    dom = et.HTML(str(soup)) 
    return dom

#funkcia, ktorá mi vráti linky na inzeráty v rámci stránky
def get_listing_url(page_url):
    dom = get_dom(page_url)
    page_link_list=dom.xpath('//a[contains(@class, "advertisement-item--content__title d-block text-truncate")]/@href')
    for page_link in page_link_list:
        listing_url.append(page_link)
        
########################################

#Atribúty, ktoré chcem dostať z inzerátu

def get_street(dom):
    try:               
        street=dom.xpath('//*[@id="map-filter-container"]/div[2]/div/div[1]/div[2]/div[2]/span/text()')
        street=street[0].strip().rstrip(',')
        
        if len(street) > 0:
            print(street)
            return street
        else:
            return "ULICA NIE JE ZADANA"
        
    except Exception as e:
        street = "ULICA NIE JE DOSTUPNA"
        print(street)

def get_rooms(dom):
    try:               
        rooms=dom.xpath('//*[@id="map-filter-container"]/div[2]/div/div[1]/div[2]/div[5]/ul/li[1]/div[2]/strong/text()')
        rooms=rooms[0].strip()
        print(rooms)
    except Exception as e:
        rooms = "IZBOVITOST NIE JE DOSTUPNA"
        print(rooms)
    return rooms

def get_sale_rent(dom):
    try:               
        sale_rent=dom.xpath('//*[@id="map-filter-container"]/div[2]/div/div[1]/div[2]/div[5]/ul/li[1]/div[1]/strong/text()') 
        sale_rent=sale_rent[0].strip()
        print(sale_rent)
    except Exception as e:
        sale_rent = "TYP VLASTNICTVA NIE JE DOSTUPNY"
        print(sale_rent)
    return sale_rent

def get_state(dom):
    try:               
        state=dom.xpath('//*[@id="map-filter-container"]/div[2]/div/div[1]/div[2]/div[5]/ul/li[1]/div[3]/strong/text()')
        state=state[0].strip()
        print(state)
    except Exception as e:
        state = "STAV NIE JE DOSTUPNY"
        print(state)
    return state

def get_usable_area(dom):
    try:               
        usable_area=dom.xpath('//*[@id="map-filter-container"]/div[2]/div/div[1]/div[2]/div[5]/ul/li[2]/div[1]/strong/text()')
        usable_area=usable_area[0].strip()
        print(usable_area)
    except Exception as e:
        usable_area = "X"
        print("VYUZITELNA PLOCHA NIE JE DOSTUPNA")
    return usable_area

def get_land_area(dom):
    try:               
        land_area=dom.xpath('//*[@id="map-filter-container"]/div[2]/div/div[1]/div[2]/div[5]/ul/li[2]/div[2]/strong/text()')
        land_area=get_land_area[0].strip()
        print(land_area)
    except Exception as e:
        land_area = "PRIESTOR POZEMKU NIE JE DOSTUPNY"
        print(land_area)
    return land_area

def get_price(dom):
    try:               
        price=dom.xpath('//*[@id="map-filter-container"]/div[2]/div/div[1]/div[2]/div[3]/div/div/div/div/div/span/text()')
        price=price[0].strip()
        print(price)
    except Exception as e:
        price = "X"
        print("CENA NIE JE DOSTUPNA")
    return price

def get_city(dom):
    try:
        city=dom.xpath('//*[@id="map-filter-container"]/div[2]/div/div[1]/div[2]/div[2]/span/a/text()')
        city=city[0].strip()
        print(city)
    except Exception as e:
        city = "MESTO NIE JE DOSTUPNE"
        print(city)
    return city