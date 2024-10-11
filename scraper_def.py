import requests
import os
import pickle
from openpyxl import load_workbook
from openpyxl import Workbook
from lxml import etree as et
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Border, Side
from fake_useragent import UserAgent
#from openpyxl.workbook import Workbook

if os.path.isfile("program_data"):
    city_list_in=open("program_data", "rb") #nacitanie miest na automaticky scraping
    city_list=pickle.load(city_list_in)
else: #vytvorenie zoznamu miest na auto scraping ak predtym neexistoval (aspon krajske mesta)
    city_list = ['Bratislava', 'Trnava', 'Nitra', 'Trenčín', 'Žilina', 'Prešov', 'Banská Bystrica', 'Košice'] 
    outputfile = open("program_data", "wb")
    pickle.dump(city_list, outputfile)
    outputfile.close()

user_agent = UserAgent()
header = {"User-Agent": user_agent.random} #hlavičky pre prehliadač (nahodne vygenerovane)
pages_url=[] #zoznam URL pre jednotlivé stránky webu
listing_url=[] #zoznam URL jednotlivých inzerátov
 
#funkcia na prehladavanie stranok a vytvorenie vysledneho suboru s datami   
def scraper(base_url, page_num, file_name, auto_mode: bool):
    pages_url.clear()
    listing_url.clear()
    
    if page_num > 1:
        for i in range (1,page_num+1): 
            page_url=base_url + "?p[page]=" + str(i)
            pages_url.append(page_url)
    else:
        pages_url.append(base_url)
        
    for page in pages_url:
        get_listing_url(page)

    current_directory = os.getcwd()
    file_path = os.path.join(current_directory, file_name) #vytvorenie suboru na zapisovanie udajov
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Vsetky data"
    else:
        wb = load_workbook(file_path)
        ws = wb["Vsetky data"]
        ws.append([]) #medzera pri dopisovani do existujuceho suboru
            
    with open(file_path, 'a', newline='') as file:
        heading=['Link', 'Mesto', 'Ulica', 'Cena', 'Izbovitosť', 'Prenájom/predaj', 'Stav', 'Úžitková plocha v m2', 'Zastavaná plocha v m2', 'EUR/m2']
        ws.append(heading)
        
        use_area_list=[] #zoznamy na vypocet pre novostavby
        eurm2_list=[]
        new_count = 0
        
        other_state=[] #zoznam na inzeraty, ktore nie su novostavby
        other_use_area_list=[]
        other_eurm2_list=[]
        other_count = 0
             
        for list_url in listing_url: 
            print("SCRAPER_DEF | URL: ", list_url)
            listing_dom=get_dom(list_url)
            
            price=get_price(listing_dom)
            try:
                price=int(''.join(filter(str.isdigit, price)))
                if price==1 or isinstance(price, str): 
                    price = "X"
            except:
                price = "X"
                
            '''if price=="X":   #Osetrenie aby nezadana cena nekazila data 
                print("\n")
                continue'''
            
            usable_area=get_usable_area(listing_dom)
            try:
                if "," in usable_area:
                    usable_area = float(usable_area.replace(',', '.').replace('m', ''))
                else:
                    usable_area=int(''.join(filter(str.isdigit, usable_area)))
            except:
                usable_area=="X"
            
            '''if usable_area=="X": #Osetrenie aby nezadana plocha bytu nekazila data
                print("\n")
                continue'''
            
            city=get_city(listing_dom)
            street=get_street(listing_dom)
            rooms=get_rooms(listing_dom)
            sale_rent=get_sale_rent(listing_dom)
            state=get_state(listing_dom)
            
            '''land_area=get_land_area(listing_dom)
            try:
                if "," in land_area:
                    land_area = float(land_area.replace(',', '.').replace('m', ''))
                else:
                    land_area=int(''.join(filter(str.isdigit, land_area)))
            except:
                land_area=="PRIESTOR POZEMKU NIE JE DOSTUPNY"'''
                       
            try:
                eurm2=price/usable_area  
            except:
                eurm2='CENA/PLOCHA V M2 NIE JE DOSTUPNA'
             
            if state=="Novostavba":
                information =[list_url, city, street, price, rooms, sale_rent, state, usable_area, eurm2] #land_area
                ws.append(information) #vypisovanie udajov inzeratov s novostavbami
                new_count += 1
            else:
                other_state.append([list_url, city, street, price, rooms, sale_rent, state, usable_area, eurm2]) #land_area
                other_count += 1
                
            if isinstance(usable_area, (int, float)) and isinstance(eurm2, (int, float)) and state=="Novostavba": #Osetrenie, aby sa vypocet vazeneho priemeru robil len vtedy, ak su oba udaje pritomne (uzitkova plocha a EUR/m2)
                use_area_list.append(usable_area)
                eurm2_list.append(eurm2)
            elif isinstance(usable_area, (int, float)) and isinstance(eurm2, (int, float)) and state!="Novostavba":
                other_use_area_list.append(usable_area)
                other_eurm2_list.append(eurm2)    
                
            print('\n') #odriadkovanie pre krajsi vypis do konzoly
          
        try:  
            new_weigh_avg = sum(x * y for x, y in zip(use_area_list, eurm2_list))/sum(use_area_list) #priemer pre novostavby
            info=['Vazeny priemer', new_weigh_avg, 'Pocet prezrenych inzeratov', new_count] 
            ws.append(info)
            ws.append([])
        except Exception as e:
            new_weigh_avg = None
            print("NIE SU NOVOSTAVBY")
        
        for insertion in other_state: #vypisanie udajov inzeratov, ktore nie su novostavby
            ws.append(insertion)
        
        try:  
            other_weigh_avg = sum(x * y for x, y in zip(other_use_area_list, other_eurm2_list))/sum(other_use_area_list) #priemer pre ostatne
            info=['Vazeny_ priemer', other_weigh_avg, 'Pocet_prezrenych inzeratov', other_count]
            ws.append(info)
            ws.append([])
        except Exception as e:
            other_weigh_avg = None
            print("NIE SU OSTATNE")
        
        wb.save(file_path)
        
        if auto_mode:
            if (new_weigh_avg is not None) and (other_weigh_avg is not None):
                avg_new_other = (new_weigh_avg, other_weigh_avg)
            elif (new_weigh_avg is not None) and (other_weigh_avg is None):
                avg_new_other = (new_weigh_avg, "X")
            elif (new_weigh_avg is None) and (other_weigh_avg is not None):
                avg_new_other = ("X", other_weigh_avg)
                
            return avg_new_other
            
        
def write_summary(file_path, city_detail_list):
    wb = load_workbook(file_path)
    ws_suhrn = wb.create_sheet("Suhrn", 0)
    ws_suhrn.title = "Suhrn"
    
    for city_detail in city_detail_list:
        tags = [city_detail["city"], "1 Izbove byty", "2 Izbove byty", "3 Izbove byty"]
        ws_suhrn.append(tags)
        new_details = [city_detail["insertion_type"]+" novostavby", city_detail["avgs"][0][0], city_detail["avgs"][1][0], city_detail["avgs"][2][0]]
        ws_suhrn.append(new_details)
        other_details = [city_detail["insertion_type"]+" ostatne", city_detail["avgs"][0][1], city_detail["avgs"][1][1], city_detail["avgs"][2][1]]
        ws_suhrn.append(other_details)
        ws_suhrn.append([])
        ws_suhrn.append([])
    
    style_summary(file_path)
    
    wb.save(file_path)
    
def style_summary(file_path):
    wb = load_workbook(file_path)
    ws_suhrn = wb["Suhrn"]

    border = Border(left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'))

    for row in ws_suhrn.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = border
        
            if isinstance(cell.value, str) and cell.value!="X":
                if cell.value.find("novostavby")!=-1 or cell.value.find("ostatne")!=-1:
                    continue
                else:
                    cell.font = Font(bold=True)
            
#funkcia na vrátenie DOM ked zadám URL  
def get_dom(the_url):
    try:
        response = requests.get(the_url, headers=header, timeout=5)
        response.raise_for_status()
    except ConnectionError as ce:
        print(f"Connection error: {ce}")
        # Handle the error, show an error message, or retry the request
        # You might want to implement a retry mechanism here
        return None
    except requests.exceptions.Timeout as e:
        print(f"Timeout error: {e}")
        # Handle the timeout error
        return None
    except requests.exceptions.RequestException as reqer:
        print(f"An error occurred during the request: {reqer}")
        # Handle other request exceptions
        return None
        
    soup = BeautifulSoup(response.text,'lxml')
    #dom = et.HTML(str(soup)) #CHANGE
    #dom_print = et.tostring(dom, pretty_print=True, encoding="utf-8").decode("utf-8")
    #print("SCRAPER_DEF | DOM: ", dom_print, "\n")
    return soup

#funkcia, ktorá mi vráti linky na inzeráty v rámci stránky
def get_listing_url(page_url):
    soup = get_dom(page_url)
    dom = et.HTML(str(soup))
    page_link_list=dom.xpath('//a[contains(@class, "advertisement-item--content__title d-block text-truncate")]/@href')
    for page_link in page_link_list:
        listing_url.append(page_link)
        
########################################

#Atribúty, ktoré chcem dostať z inzerátu. Keď sa zmení DOM (rozloženie stránky), treba chcené atribúty nanovo lokalizovať.

def get_street(dom): #ulica, na ktorej je nehnuteľnosť
    try:               
        #street=dom.xpath('//*[@id="__next"]/div/main/div/div/div/div[2]/div[2]/div[1]/div/div[1]/div/div[2]/div/div[1]/p')
        street=dom.select_one('#__next > div > main > div > div > div > div.MuiContainer-root.MuiContainer-maxWidthLg.css-1qsxih2 > div.MuiGrid-root.MuiGrid-container.MuiGrid-spacing-xs-2.css-isbt42 > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-lg-8.css-1mrfx1k > div > div.MuiBox-root.css-19idom > div > div:nth-child(3) > div > div.MuiStack-root.css-1r9kwv0 > p')
        print(street)
        street = street.text.strip().split(',')[0]
        
        if len(street) > 0:
            return street
        else:
            return "ULICA NIE JE ZADANA"
        
    except Exception as e:
        print("CHYBA PRI ULICI: ", e)
        street = "ULICA NIE JE DOSTUPNA"

def get_rooms(dom): #izbovitosť nehnuteľnosti (byt)
    try:               
        #rooms=dom.xpath('//*[@id="__next"]/div/main/div/div/div/div[2]/div[2]/div[1]/div/div[1]/div/div[2]/div/div[2]/div/text()[2]')
        rooms=dom.select_one('#__next > div > main > div > div > div > div.MuiContainer-root.MuiContainer-maxWidthLg.css-1qsxih2 > div.MuiGrid-root.MuiGrid-container.MuiGrid-spacing-xs-2.css-isbt42 > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-lg-8.css-1mrfx1k > div > div.MuiBox-root.css-19idom > div > div:nth-child(3) > div > div.MuiBox-root.css-70qvj9 > div > :first-child::text')
        print(rooms)
        rooms=rooms.text.strip()
    except Exception as e:
        rooms = "IZBOVITOST NIE JE DOSTUPNA"
        print("CHYBA PRI IZBOVITOSTI: ", e)
    return rooms

def get_sale_rent(dom): #ci je nehnuteľnosť na prenájom alebo predaj
    try:               
        #sale_rent=dom.xpath('//*[@id="__next"]/div/main/div/div/div/div[2]/div[2]/div[1]/div/div[3]/div[1]/div/h2')
        sale_rent=dom.select_one('#__next > div > main > div > div > div > div.MuiContainer-root.MuiContainer-maxWidthLg.css-1qsxih2 > div.MuiGrid-root.MuiGrid-container.MuiGrid-spacing-xs-2.css-isbt42 > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-lg-8.css-1mrfx1k > div > div:nth-child(3) > div.MuiBox-root.css-i3pbo > div > h2')
        print(sale_rent)
        sale_rent=sale_rent.split(' ')[-1]
    except Exception as e:
        sale_rent = "TYP VLASTNICTVA NIE JE DOSTUPNY"
        print("CHYBA PRI VLASTNICTVE: ", e)
    return sale_rent

def get_state(dom): #či je nehnuteľnosť novostavba alebo iný stav
    try:               
        #state=dom.xpath('//*[@id="__next"]/div/main/div/div/div/div[2]/div[2]/div[1]/div/div[1]/div/div[2]/div/div[2]/div/text()[8]')
        state=dom.select_one('#__next > div > main > div > div > div > div.MuiContainer-root.MuiContainer-maxWidthLg.css-1qsxih2 > div.MuiGrid-root.MuiGrid-container.MuiGrid-spacing-xs-2.css-isbt42 > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-lg-8.css-1mrfx1k > div > div:nth-child(3) > div.MuiBox-root.css-i3pbo > div > div > div:nth-child(2) > div > div > p')
        print(state)
        state=state.text.strip()
    except Exception as e:
        state = "STAV NIE JE DOSTUPNY"
        print("CHYBA PRI STAVE: ", e)
    return state

def get_usable_area(dom): #úžitková plocha nehnuteľnosti, interiér
    try:               
        #usable_area=dom.xpath('//*[@id="__next"]/div/main/div/div/div/div[2]/div[2]/div[1]/div/div[1]/div/div[2]/div/div[2]/div/text()[5]')
        usable_area=dom.select_one('#__next > div > main > div > div > div > div.MuiContainer-root.MuiContainer-maxWidthLg.css-1qsxih2 > div.MuiGrid-root.MuiGrid-container.MuiGrid-spacing-xs-2.css-isbt42 > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-lg-8.css-1mrfx1k > div > div.MuiBox-root.css-19idom > div > div:nth-child(3) > div > div.MuiBox-root.css-70qvj9 > div > :nth-child(2)::text')
        print(usable_area)
        usable_area=usable_area.text.strip()
    except Exception as e:
        usable_area = "X"
        print("CHYBA PRI VYUZITELNEJ PLOCHE: ", e)
    return usable_area

'''def get_land_area(dom): #úžitková plocha nehnuteľnosti, exteriér | PO ZMENE STRÁNKY ZLOŽITEJŠIE ZÍSKAŤ
    try:               
        land_area=dom.xpath('') 
        land_area=get_land_area[0].strip()
        print(land_area)
    except Exception as e:
        land_area = "PRIESTOR POZEMKU NIE JE DOSTUPNY"
        print(land_area)
    return land_area'''

def get_price(dom): #kúpna cena/cena prenájmu
    try:               
        #price_elem=dom.xpath('//*[@id="__next"]/div/main/div/div/div/div[2]/div[2]/div[1]/div/div[1]/div/div[3]/div/p[1]')
        price=dom.select_one('#__next > div > main > div > div > div > div.MuiContainer-root.MuiContainer-maxWidthLg.css-1qsxih2 > div.MuiGrid-root.MuiGrid-container.MuiGrid-spacing-xs-2.css-isbt42 > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-lg-8.css-1mrfx1k > div > div.MuiBox-root.css-19idom > div > div:nth-child(4) > div > p.MuiTypography-root.MuiTypography-h3.css-wupcfx')
        print(price)
        
        price = price.text.strip()
        
    except Exception as e:
        print("Chyba pri ziskavani ceny: ", e)
        price = "X"
    return price

def get_city(dom): #mesto, kde sa nehnuteľnosť nachádza
    try:
        #city=dom.xpath('//*[@id="__next"]/div/main/div/div/div/div[2]/div[2]/div[1]/div/div[1]/div/div[2]/div/div[1]/p')
        city=dom.select_one('#__next > div > main > div > div > div > div.MuiContainer-root.MuiContainer-maxWidthLg.css-1qsxih2 > div.MuiGrid-root.MuiGrid-container.MuiGrid-spacing-xs-2.css-isbt42 > div.MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-12.MuiGrid-grid-lg-8.css-1mrfx1k > div > div.MuiBox-root.css-19idom > div > div:nth-child(3) > div > div.MuiStack-root.css-1r9kwv0 > p')
        print(city)
        city = city.text.strip().split(',')[1:]
        city=','.join(city)
    except Exception as e:
        city = "MESTO NIE JE DOSTUPNE"
        print("CHYBA PRI MESTE: ", e)
    return city